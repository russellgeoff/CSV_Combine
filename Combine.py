"""
Description:
Combine multiple CSV data files into one Excel document
"""

import os
import glob
import csv
import openpyxl
import natsort

__author__ = 'Geoff Russell'
__version__ = '1.0'

print("CSV Combine version: %s \n"
      "Created by %s \n\n" % (__version__, __author__))
filename = raw_input("What do you want the merged file to be called? ")
while filename == "":
    print("ERROR: Filename cannot be blank!")
    filename = raw_input("What do you want the merged file to be called? ")

#Sets up workbook & adds data summary sheet
summary_row = 1
summary_ws_data = {
    'title': 'Data Summary',
    'col1_title': 'Sample',
    'col1_formula': "='%s'!$B$1",
    'col2_title': 'Removal',
    'col2_formula': "=ABS(MIN('%s'!C:C))",
    'col3_title': 'Notes',
    'col3_formula': "='%s'!$B$2"
}

wb = openpyxl.Workbook()
ws_summary = wb.active
ws_summary.title = summary_ws_data['title']
ws_summary.cell(row=summary_row, column=1).value = summary_ws_data['col1_title']
ws_summary.cell(row=summary_row, column=2).value = summary_ws_data['col2_title']
ws_summary.cell(row=summary_row, column=3).value = summary_ws_data['col3_title']

#Gets list of all CSV files in the current directory and sorts them intelligently
files = glob.glob(os.path.join('.', '*.csv'))
sortedFiles = natsort.natsorted(files)

#Puts each CSV file into a new sheet
for csv_file in sortedFiles:
    head, tail = os.path.split(csv_file)
    ws = wb.create_sheet()
    ws.title = tail
    with open(csv_file, 'rb') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, val in enumerate(row):
                cell = ws.cell(row=r+1, column=c+1)
                try:
                    cell.value = float(val)  # Tries to convert all numbers to floats
                except ValueError:
                    try:
                        cell.value = str(val)  # Tries to submit the string value
                    except UnicodeDecodeError:
                        try:
                            print("Error decoding %s in cell (row = %i, column = %i)\n"
                                  "Trying a different encoding standard (Unicode ISO-8859-1)"
                                  % (val, r+1, c+1))
                            cell.value = unicode(val, "ISO-8859-1")
                            print("Different encoding standard worked!")
                        except:
                            print("Decoding failed\nThis cell will be blank")
                except:
                    print("Other Error\nError occured in cell "
                          "(row = %i, column = %i). The value was %s"
                          % (r, c, val))
                    print val
    # Summary Worksheet Information
    summary_row += 1
    # ws_summary.cell(row=summary_row, column=1).value = ws.title
    ws_summary.cell(row=summary_row, column=1).value = summary_ws_data['col1_formula'] % ws.title

    ws_summary.cell(row=summary_row, column=2).value = summary_ws_data['col2_formula'] % ws.title

    ws_summary.cell(row=summary_row, column=3).value = summary_ws_data['col3_formula'] % ws.title

    
    # ws_summary.cell(row=summary_row, column=3).value = '=MAX(%s!C:C)' % ws.title
    # ws_summary.cell(row=summary_row, column=3).value = '=MAX(%s!C:C))' % ws.title
    print('Added %s' % tail)

wb.save(filename + '.xlsx')
print('Wrote ' + filename)
